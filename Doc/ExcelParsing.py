# -*- coding: utf-8 -*-
"""
Created on Mon Sep 21 14:06:03 2020

@author: peezflay
"""
#pip install openpyxl
#https://pypi.org/project/openpyxl/#description

import openpyxl as excel
import pprint
from openpyxl.utils import get_column_letter
from array import *
import datetime

wb = excel.load_workbook('test.xlsx')
ProductInfo = wb['Sheet1']
ProductRaw = wb['Sheet2']
FinalReport = wb['Sheet3']

ProductStruct = {}
ProductStructs = {}


for row in range(3, ProductInfo.max_row+1):
    current_row = []
    if(ProductInfo['M'+str(row)].value != "Paid"):
        for col in range(1, ProductInfo.max_column+1):
            current_row.append(ProductInfo.cell(row=row,column=col).value)
        ProductRaw.append(current_row)
    else:
        if(ProductInfo['C'+str(row)].value == "Payment Fee" or ProductInfo['C'+str(row)].value == "Promotional Charges Vouchers"):
            for col in range(1, ProductInfo.max_column+1):
                current_row.append(ProductInfo.cell(row=row,column=col).value)
                Temp_FeeName = ProductInfo['C'+str(row)].value
            #nPaymentFeeExistFlag = 0
            for Sheet2_row in range(1, ProductRaw.max_row+1):
                if(ProductRaw['C'+str(Sheet2_row)].value == Temp_FeeName):
                    Temp_FVar = ProductRaw['H'+str(Sheet2_row)].value or 0
                    ProductRaw['H'+str(Sheet2_row)].value = float(Temp_FVar) + ProductInfo['H'+str(row)].value
                    break
            else:
                ProductRaw.append(current_row)
        elif("Shipping Fee" in ProductInfo['C'+str(row)].value):      
            for col in range(1, ProductInfo.max_column+1):
                if(col != 3):
                    current_row.append(ProductInfo.cell(row=row,column=col).value)
                else:
                    current_row.append("Shipping Fee Overall")
            for Sheet2_row in range(1, ProductRaw.max_row+1):
                if(ProductRaw['C'+str(Sheet2_row)].value == "Shipping Fee Overall"):
                    Temp_FVar = ProductRaw['H'+str(Sheet2_row)].value or 0
                    ProductRaw['H'+str(Sheet2_row)].value = float(Temp_FVar) + ProductInfo['H'+str(row)].value
                    Temp_nVar = ProductRaw['V'+str(Sheet2_row)].value or 0
                    ProductRaw['V'+str(Sheet2_row)].value = Temp_nVar + 1                    
                    break
            else:
                current_row.append(1);
                ProductRaw.append(current_row)
        elif(ProductInfo['C'+str(row)].value == "Item Price Credit"):
            for col in range(1, ProductInfo.max_column+1):
                current_row.append(ProductInfo.cell(row=row,column=col).value)
                Temp_SellerSKU = ProductInfo['F'+str(row)].value.split("-q")
            for Sheet2_row in range(1, ProductRaw.max_row+1):
                if(ProductRaw['C'+str(Sheet2_row)].value == "Item Price Credit" and ProductRaw['F'+str(Sheet2_row)].value == Temp_SellerSKU[0]):
                    Temp_FVar = ProductRaw['H'+str(Sheet2_row)].value or 0
                    ProductRaw['H'+str(Sheet2_row)].value = float(Temp_FVar) + ProductInfo['H'+str(row)].value
                    Temp_nVar = ProductRaw['V'+str(Sheet2_row)].value or 0
                    if(len(Temp_SellerSKU) > 1):
                        ProductRaw['V'+str(Sheet2_row)].value = Temp_nVar + int(Temp_SellerSKU[1])
                    else:
                        ProductRaw['V'+str(Sheet2_row)].value = Temp_nVar + 1 
                    break
            else:
                current_row.append(1);
                ProductRaw.append(current_row)                
        else:
            for col in range(1, ProductInfo.max_column+1):
                current_row.append(ProductInfo.cell(row=row,column=col).value)
            ProductRaw.append(current_row)
 
FinalReportColumnName = ["#", "Type *", "Customer *", "Transaction No*", "Date *", "Service Date", "Terms +", "Due Date / Expiry Date/Delivery Date +", "Reference", "Export No.", "Export Country", "Special Scheme No.", "Contact Person", "Address Line 1", "Address Line 2", "Address Line 3", "Address Line 4", "Currency Code *", "Currency Exchange Rate +", "Item/ Account Name *", "Item / Account Description", "Quantity *", "Unit Price *", "Discount", "Tax Code", "Tax Rate", "Tariff Code", "Job No", "Transaction Note", "Delivery Note", "Journal Memo"];
if(FinalReport['A2'] != "#"):
    FinalReport.append(FinalReportColumnName)


CurrentDate = datetime.datetime.now()   
TransactionNo = "LZD"+CurrentDate.strftime("%Y%m%d")
DueDate = CurrentDate + datetime.timedelta(days=30)

FinalReportColumn = ["", "Invoice", "Lazada", TransactionNo, CurrentDate.strftime("%x"), "", 30, DueDate.strftime("%x"), "", "", "", "", "", "", "", "", "", "SGD", "", "Item/Account Name", "Item/Account Description", "Quantity", "Unit Price"]
for row in range(2, ProductRaw.max_row+1) :
    if(ProductRaw['C'+str(row)].value == "Item Price Credit"):
        FinalReportColumn[19] = ProductRaw['F'+str(row)].value
        FinalReportColumn[20] = ProductRaw['E'+str(row)].value
    elif(ProductRaw['C'+str(row)].value == "Payment Fee"):
        FinalReportColumn[19] = "Payment Fee"
        FinalReportColumn[20] = ProductRaw['C'+str(row)].value
    elif(ProductRaw['C'+str(row)].value == "Shipping Fee Overall"):
        FinalReportColumn[19] = "Shipping Fee"
        FinalReportColumn[20] = ProductRaw['C'+str(row)].value
    elif("Promotional Charges" in ProductRaw['C'+str(row)].value):
        FinalReportColumn[19] = "Promotion Charges"
        FinalReportColumn[20] = ProductRaw['C'+str(row)].value
    Temp_Quantity = ProductRaw['V'+str(row)].value or 0
    if(Temp_Quantity < 1):
        FinalReportColumn[21] = 1
    else:
        FinalReportColumn[21] = Temp_Quantity
    FinalReportColumn[22] = ProductRaw['H'+str(row)].value
    FinalReport.append(FinalReportColumn)
    
    
    
    
    

    
    
    
#for row in range(2, ProductInfo.max_row+1):
#    Product = ProductInfo['B'+str(row)].value
#    ProductStruct.setdefault(Product, {'Description':'0'})
#    ProductStruct[Product]['Description']  = ProductInfo['D'+str(row)].value
#    
#pprint.pprint(ProductStruct)
#
#for machine in ProductStruct:
#    if machine not in ProductStructs:
#        ProductStructs.setdefault(machine, {'Description':'0'})
#        ProductStructs[machine]['Description'] = ProductStruct[machine]['Description']
#    
#        
#
#Temp = 1
#for machine in ProductStruct:
#    ProductProfits['A'+str(Temp)].value =machine
#    ProductProfits['B'+str(Temp)].value =ProductStruct[machine]["Description"]
#    Temp+=1
#    
    
wb.save('test.xlsx')
    