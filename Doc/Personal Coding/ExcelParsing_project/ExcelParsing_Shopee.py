import openpyxl as excel
import pprint
from openpyxl.utils import get_column_letter
from openpyxl.styles import Color, PatternFill, Font, Border
from array import *
import datetime

wb = excel.load_workbook('Shopee.xlsx')

try:
    IncomeTab = wb["Income"]
    OrderTab = wb["orders"]
except KeyError:
    print("Creating Sheet which does not exist.....")
    
sheet_name = wb.sheetnames
for SN in sheet_name:
    if SN == "Income" or SN == "orders":
        continue
    RemoveSheet = wb[SN]
    wb.remove(RemoveSheet)
    
RawSheet = wb.create_sheet(title="Sheet1")
FinalReport = wb.create_sheet(title="Sheet2")
RawSheetAppendList = []
#RawSheetListDict = [{"OrderID": "", "Quantity": "", "DealPrice": "", "SumOfDealPrice": "", "UnitPrice": ""} for k in range(IncomeTab.max_row+1)]
RawSheetListDict = [{"OrderID": "", "Quantity": "", "DealPrice": "", "SumOfDealPrice": "", "UnitPrice": ""}]
RedFill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')

RawSheet.append(["Order ID", "total released amount($)", "SKU reference No.", "Quantity", "Deal price", "Sum of deal price for this Order ID", "Unit Price"])

for row in range(1, IncomeTab.max_row+1):
    Temp_Var = IncomeTab.cell(row=row, column=1).value or ""
    if type(Temp_Var) == str and "Sequence" in Temp_Var:
        IncomeTabStartRow = row
        
for row in range(1, OrderTab.max_row+1):
    Temp_Var = OrderTab.cell(row=row, column=1).value or ""
    if type(Temp_Var) == str and "Order" in Temp_Var:
        OrderTabStartRow = row
        
for IncomeTabRow in range(IncomeTabStartRow+1, IncomeTab.max_row+1):
    Temp_OrderID = IncomeTab["B"+str(IncomeTabRow)].value or ""
    nFound = 0
    for OrderTabRow in range(OrderTabStartRow+1, OrderTab.max_row+1):
        if len(Temp_OrderID) > 1:
            if Temp_OrderID  in OrderTab["A"+str(OrderTabRow)].value:
                TempSKURef = OrderTab["M"+str(OrderTabRow)].value or ""
                if len(TempSKURef) < 1:
                    TempSKURef = OrderTab["K"+str(OrderTabRow)].value or "SKU Reference No. Empty"
                Temp_SellerSKU = TempSKURef.split("-q")
                if(len(Temp_SellerSKU) > 1):
                    Temp_Quantity = OrderTab["Q"+str(OrderTabRow)].value * int(Temp_SellerSKU[1])
                else:
                    Temp_Quantity = OrderTab["Q"+str(OrderTabRow)].value
                RawSheetAppendList.extend([Temp_OrderID, IncomeTab["R"+str(IncomeTabRow)].value, Temp_SellerSKU[0], Temp_Quantity, OrderTab["P"+str(OrderTabRow)].value])
                RawSheet.append(RawSheetAppendList)
                RawSheetAppendList.clear()
                nFound += 1
            if nFound == 0 and OrderTabRow == OrderTab.max_row:
                RawSheetAppendList.append(Temp_OrderID)
                RawSheet.append(RawSheetAppendList)
                RawSheetAppendList.clear()
                RawSheet["A"+str(OrderTabRow-1)].fill = RedFill

for RawSheetRow in range(2, RawSheet.max_row+1):
    Temp_OrderID = RawSheet["A"+str(RawSheetRow)].value or ""
    Temp_Quantiy = RawSheet["D"+str(RawSheetRow)].value or 0
    Temp_DealPrice = RawSheet["E"+str(RawSheetRow)].value or 0
    nFound = 0
    for RawSheetDictLoop in range(len(RawSheetListDict)):
        if RawSheetListDict[RawSheetDictLoop]["OrderID"] == Temp_OrderID:
            RawSheetListDict[RawSheetDictLoop]["Quantity"] += Temp_Quantiy
            RawSheetListDict[RawSheetDictLoop]["DealPrice"] += Temp_DealPrice
            RawSheetListDict[RawSheetDictLoop]["SumOfDealPrice"] +=  Temp_Quantiy*Temp_DealPrice
            nFound += 1
            
    if nFound == 0:
        RawSheetListDict.append({"OrderID": Temp_OrderID, "Quantity": Temp_Quantiy, "DealPrice": Temp_DealPrice, "SumOfDealPrice": Temp_Quantiy*Temp_DealPrice, "UnitPrice": ""})
    
for item in RawSheetListDict:
    print(item)

for RawSheetRow in range(2, RawSheet.max_row+1):
    Temp_OrderID = RawSheet["A"+str(RawSheetRow)].value or ""
    Temp_Quantiy = RawSheet["D"+str(RawSheetRow)].value or 0
    Temp_DealPrice = RawSheet["E"+str(RawSheetRow)].value or 0
    Temp_TotAmtRel = RawSheet["B"+str(RawSheetRow)].value or 0
    if Temp_Quantiy < 1:
        continue
    for RawSheetDictLoop in range(len(RawSheetListDict)):
        if RawSheetListDict[RawSheetDictLoop]["OrderID"] == Temp_OrderID:
            Temp_Tot_Quantity = RawSheetListDict[RawSheetDictLoop]["Quantity"]
            Temp_Tot_DealPrice = RawSheetListDict[RawSheetDictLoop]["DealPrice"]
            Temp_Tot_SumOfDealPrice = RawSheetListDict[RawSheetDictLoop]["SumOfDealPrice"]
    RawSheet['F'+str(RawSheetRow)].value = Temp_Tot_SumOfDealPrice
    RawSheet['G'+str(RawSheetRow)].value = (((Temp_DealPrice*Temp_Quantiy)/Temp_Tot_SumOfDealPrice)*Temp_TotAmtRel)/Temp_Quantiy  #This Last divide quantity is solo quantity or total quantity


FinalReportColumnName = ["#", "Type *", "Customer *", "Transaction No*", "Date *", "Service Date", "Terms +", "Due Date / Expiry Date/Delivery Date +", "Reference", "Export No.", "Export Country", "Special Scheme No.", "Contact Person", "Address Line 1", "Address Line 2", "Address Line 3", "Address Line 4", "Currency Code *", "Currency Exchange Rate +", "Item/ Account Name *", "Item / Account Description", "Quantity *", "Unit Price *", "Discount", "Tax Code", "Tax Rate", "Tariff Code", "Job No", "Transaction Note", "Delivery Note", "Journal Memo"];
if(FinalReport['A2'] != "#"):
    FinalReport.append(FinalReportColumnName)


CurrentDate = datetime.datetime.now()   
TransactionNo = "SHP"+CurrentDate.strftime("%Y%m%d")
DueDate = CurrentDate + datetime.timedelta(days=30)

FinalReportColumn = ["", "Invoice", "Shopee", TransactionNo, CurrentDate.strftime("%x"), "", 30, DueDate.strftime("%x"), "", "", "", "", "", "", "", "", "", "SGD", "", "Item/Account Name", "", "Quantity", "Unit Price"]
for RawSheetRow in range(2, RawSheet.max_row+1) :
    Temp_Quantiy = RawSheet["D"+str(RawSheetRow)].value or 0
    if Temp_Quantiy < 1:
        continue
    FinalReportColumn[19] = RawSheet['C'+str(RawSheetRow)].value
    Temp_Quantity = RawSheet['D'+str(RawSheetRow)].value or 0
    if(Temp_Quantity > 0):
        FinalReportColumn[21] = RawSheet['D'+str(RawSheetRow)].value
    FinalReportColumn[22] = RawSheet['G'+str(RawSheetRow)].value
    FinalReport.append(FinalReportColumn)
    
wb.save('Shopee.xlsx')